from openpyxl import load_workbook


def read_excel_as_dict_of_dicts(file_path, sheet_name, key_column="TCName"):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in sheet[1]]  # first row as headers
    key_index = headers.index(key_column)

    data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[key_index]
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i != key_index}
        data[key] = row_dict

    return data
